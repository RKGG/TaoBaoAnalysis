__author__ = "* * treaser 2018 * * "

from selenium import webdriver
import time
import random
from selenium.webdriver.common.action_chains import ActionChains
import pymysql
import openpyxl
import requests
from bs4 import BeautifulSoup


def show_info():
    info = "Version-1.1 \
    Time-2018.5.30 \
    two bugs: \
    1. verify the slide bar. \
    2. even if the slide bar be verified, the click to the button seems wrong."
    print(info)


def taobao_request_login():
    params = {"TPL_username": "treasershere", "TPL_password": "1996818321lusio"}
    taobao = requests.post("https://login.taobao.com/member/login.jhtml?spm=a21bo.2017.201864-2.d1.5af911d9a1ABPg&f=top&redirectURL=http%3A%2F%2Fwww.taobao.com%2F",data=params)
    print("Cookie is set to: ")
    print(taobao.cookies.get_dict())
    print("--------------------------------------")
    print("Going to list_bought_items page...")
    list_bought_items = requests.get("https://buyertrade.taobao.com/trade/itemlist/list_bought_items.htm?spm=a1z02.1.a2109.d1000368.584d782dIjzjZv&nekot=1470211439694", cookies=taobao.cookies)
    # print(list_bought_items.content)
    print("--------------------------------------")
    bs = BeautifulSoup(list_bought_items.content, "html.parser")   # list_bought_items.text
    print(bs.prettify())


class TaoBao:

    def __init__(self, username, password):
        self.username = username
        self.password = password
        self.url = "https://login.taobao.com/" # http://www.taobao.com
        self.driver = webdriver.Firefox()  # Firefox
        self.service_args = [
            # '--proxy=218.241.30.187:8123',
            # '--proxy-type=http',
        ]
        self.loginHeaders = {
            'Host': 'login.taobao.com',
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64; rv:35.0) Gecko/20100101 Firefox/35.0',
            'Referer': 'https://login.taobao.com/member/login.jhtml',
            'Content-Type': 'application/x-www-form-urlencoded',
            'Connection': 'Keep-Alive'
        }

    def send_keys_slowly(self, web_element, key):
        for i in range(1, len(key) + 1):
            web_element.send_keys(key[i - 1])
            time.sleep(0.1)

    def find_and_perform_slide_bar(self):
        slide_bar = self.driver.find_element_by_id("nc_1_n1z")
        slide_action = ActionChains(self.driver)
        slide_action.click_and_hold(slide_bar).perform()
        slide_action.reset_actions()
        # slide_action.reset_actions()
        # slide_action.move_by_offset(400, 0).perform()
        valper = 100 * random.random()  # valper = 100 * random.random()
        total = 0
        while total < 260:  # 400 其实好像是258
            try:
                slide_action.move_by_offset(valper, 0).perform()  # 平行移动鼠标
                total += valper
                valper = 100 * random.random()
            except:
                time.sleep(2)
                print("in find_and_do_slidebar exception")
                break
            time.sleep(0.2)  # 等待停顿时间
        try:
            direction = self.driver.find_elements_by_class_name("nc-lang-cnt")
            if direction.text == "验证通过":
                print("Verify Pass!!!")
            else:
                print("Still not verified...")
        except:
            print("Still not verified...")
            pass

    def common_click(self, element, sleeptime=1):  # ATN: when call this method, first para is element....
        actions = ActionChains(self.driver)
        actions.move_to_element(element)
        actions.click(element)
        actions.perform()
        time.sleep(sleeptime)

    def login_by_hand(self, maxtimes=1):
        starttime = time.time()
        self.driver.implicitly_wait(3)
        self.driver.get(self.url)  # https://tieba.baidu.com/index.html?nu_token=4b8a703069686a352a64346a246472652cb0
        # pswd_login = self.driver.find_element_by_link_text("密码登录")
        # self.common_click(pswd_login)
        time.sleep(1)
        cmd = input("=_=...\n")
        if cmd == "go":
            return 0
        else:
            return -1

    def login(self, maxtimes=1):
        starttime = time.time()
        self.driver.implicitly_wait(3)
        self.driver.get(self.url)  # https://tieba.baidu.com/index.html?nu_token=4b8a703069686a352a64346a246472652cb0
        #print("cookies before login: ")
        #print(self.driver.get_cookies())
        time.sleep(2) #
        # pre_login_button = self.driver.find_element_by_link_text("亲，请登录")
        # print(pre_login_button)
        # self.common_click(pre_login_button)
        # time.sleep(1)
        pswd_login = self.driver.find_element_by_link_text("密码登录")
        self.common_click(pswd_login)
        time.sleep(1)
        i = 0
        while i < maxtimes:
            print("Try to login for the " + str(i + 1) + " th time.")
            i += 1
            username_txt = self.driver.find_element_by_id("TPL_username_1")
            username_txt.clear()
            self.send_keys_slowly(username_txt, self.username)
            userpswd_txt = self.driver.find_element_by_id("TPL_password_1")
            userpswd_txt.clear()
            self.common_click(userpswd_txt)
            self.send_keys_slowly(userpswd_txt, self.password)
            time.sleep(2)
            # When finishing tpying your username and password, there are two situations:
            try:  # 1 there being a slide bar
                self.find_and_perform_slide_bar()
            except:
                print("Maybe sth. wrong with slide bar..")
                try:  # 2 there being no slide bar, but may have a refresh button to generate a slide bar.
                    refresh_button = self.driver.find_element_by_link_text("刷新")
                    self.common_click(refresh_button)
                    print("-Because we need to press the refresh button.")
                    time.sleep(1)
                    self.find_and_perform_slide_bar()
                except:  # no refreshing button & no slide bar, we can just easily submit !
                    print("-Because no slide bar to verify!")

            print("Try to submit...")
            login_button = self.driver.find_element_by_id("J_SubmitStatic")
            self.common_click(login_button)
            print("Already clicked the sign-in button.")

            try:
                self.driver.find_element_by_id("TPL_username_1")
                print("Login failed once!")
            except:
                print("Login successfully!")
                endtime = time.time()
                print("Login time cost:" + str(endtime - starttime))
                #print("cookies after login: ")
                #print(self.driver.get_cookies())
                return 0  # login succeed
        return -1   # login fail

    def perform(self, bought_dates, item_names, bought_prices):
        starttime = time.time()

        self.common_click(self.driver.find_element_by_link_text("我的淘宝"))
        time.sleep(2)
        # driver.find_element_by_link_text("已买到的宝贝").click()
        self.driver.find_element_by_id("bought").click()
        # common_click(driver, driver.find_element_by_link_text("已买到的宝贝"))
        time.sleep(2)
        bought_dates_obj = self.driver.find_elements_by_class_name("bought-wrapper-mod__create-time___yNWVS")
        # bought_dates = []
        for bought_date_obj in bought_dates_obj:
            bought_dates.append(bought_date_obj.text)
            print("date:" + bought_date_obj.text)

        # item_names = []
        for i in range(1, 50):
            try:
                item_names_obj = self.driver.find_element_by_xpath(
                    "/html/body/div[2]/div/div[1]/div[1]/div[3]/div/div[" + str(i + 3) + "]/div/table/tbody[2]/tr/td[1]/div/div[2]/p[1]/a/span[2]")
                item_name = item_names_obj.text
                print("name: " + item_name)
                item_names.append(item_name)
            except:
                break

        # bought_prices = []
        for i in range(1, 50):
            try:
                bought_price_obj = self.driver.find_element_by_xpath("/html/body/div[2]/div/div[1]/div[1]/div[3]/div/div[" + str(i + 3) + "]/div/table/tbody[2]/tr[1]/td[5]/div/div[1]/p/strong/span[2]")
                bought_price = bought_price_obj.text
                print("price: " + bought_price)
                bought_prices.append(bought_price)
            except:
                break

        for i in range(len(bought_dates)):
            print(bought_dates[i] + " : " + item_names[i] + " " + bought_prices[i] + " Yuan.")

        endtime = time.time()
        print("Peform time cost:" + str(endtime - starttime))
        data = []
        for i in range(len(bought_dates)):
            data[i] = []
            data[i][0] = bought_dates[i]
            data[i][0] = item_names[i]
            data[i][0] = bought_prices[i]


class MySqlDBStore:
    # To use Mysql you must FIRST CREATE the the DATABASE and the TABLE you would like to store data in !!!
    # In my project I use the TABLE taobaoanalysis in DATABASE usualdata to store my data.

    def __init__(self, user, pswd):
        self.user = user
        self.pswd = pswd

    def login(self):
        self.connect = pymysql.connect(host='127.0.0.1', unix_socket='/tmp/mysql.sock', user=self.user, passwd=self.pswd,
                                  db='mysql', charset="utf8mb4")   #, autocommit=True
        # self.cursor = self.connect.cursor()

    def execute(self, cmd):
        self.connect.cursor().execute(cmd)
        print(self.connect.cursor().fetchall())

    def store(self,bought_dates,item_names,bought_prices):
        records = []
        for i in range(len(bought_dates)):
            record = []
            record.append(str(bought_dates[i]))
            record.append(str(item_names[i]))
            record.append(str(bought_prices[i]))
            records.append(record)
        self.connect.cursor().execute("use usualdata")  # must use self.connect.cursor() instead of self.cursor !!
        with self.connect.cursor() as cursor:
            #  cursor.executemany("insert into test(prop, val) values (%s, %s)", vals)
            cmd = "insert into taobaoanalysis(id,bought_date,item_name,bought_price) values(NULL,%s,%s,%s)"
            cursor.executemany(cmd, records)
            self.connect.commit()


class Store:

    def __init__(self, form = "txt"):
        self.form = form

    def txtstore(self, data, name="store/taobao.txt"):
        f = open(name, 'r')
        line_num = 0
        for line in f:
            line_num+=1
        print("lines: " + str(line_num) + "\n")
        f.close()
        f = open(name, 'w')
        f.write("Revise time: "+str(time.asctime(time.localtime(time.time()))) + "\n")
        for i in range(len(data)):
            f.write(str(i+1))
            f.write("==")
            f.write(data[i][0])
            f.write("==")
            f.write(data[i][1])
            f.write("==")
            f.write(data[i][2])
            f.write("\n")
        f.close()

    def excelstore(self, data, name="store/taobao.xlsx"):
        xlsx = openpyxl.Workbook()
        newsheet = xlsx.active
        for i in range(len((data))):
            newsheet.cell(row=i+1, column=1, value=data[i][0])  # .value = newss_date
            newsheet.cell(row=i+1, column=2, value=data[i][1])
            newsheet.cell(row=i+1, column=3, value=data[i][2])
            # print(news_date, news_content)
        xlsx.save(name)


if __name__ == "__main__":
    data = [["2018-04-27","中国电信官方旗舰店浙江手机充值100元电信话费直充快充电信充值","99.80"],["2018-05-27","8MHZ晶振 HC-49SMD 2脚贴片 假贴 8M","61.06"]]
    store = Store()
    store.txtstore(data)
    store.excelstore(data)
    taobao_request_login()